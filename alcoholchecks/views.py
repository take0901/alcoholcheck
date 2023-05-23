from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from users.models import User
from django.http import Http404
from .models import Info
from .forms import InfoForm
import datetime
import io
import openpyxl as xl
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import zipfile
from django.http import JsonResponse

def index(request):
    #ホームページ
    check = ""
    download = ""
    if str(request.user) == "alcohol_admin":
        check = "チェックする"
        download = "ダウンロード"
    context = {'check':check, "download":download}
    return render(request, 'alcoholchecks/index.html', context)

@login_required
def infos(request,user_id):
    user = User.objects.get(id=user_id)
    infos = user.info_set.order_by("-date_added")
    delete = ""
    check = ""
    if user != request.user and str(request.user) != "alcohol_admin":
        raise Http404
    if str(request.user) == "alcohol_admin":
        delete = "削除"
        check = "チェックする"
    context = {'infos':infos, 'delete':delete, 'check':check, 'user':user}
    return render(request, 'alcoholchecks/infos.html', context)

@login_required
def new_info(request, user_id):
    user = User.objects.get(id=user_id)
    if user != request.user and str(request.user) != "alcohol_admin":
        raise Http404
        
    if request.method != 'POST':
        #フォームを生成
        form = InfoForm(initial={'carnumber':user.carnumber, 'alcohol':"アルコール検知: 0.00mg"})
        form.fields['carnumber'].choices = list(User.objects.values_list("carnumber", 
                                                            'carnumber').distinct())
    else:
        #POSTで送信されたデータを処理
        form = InfoForm(data=request.POST)
        if form.is_valid():
            new_info = form.save(commit=False)
            new_info.owner = user
            new_info.save()
            return redirect('alcoholchecks:infos', user_id=user_id)

    #空のフォームを表示させる
    context = {'form':form, 'user':user}
    return render(request, 'alcoholchecks/new_info.html', context)

@login_required
def check(request):
    if str(request.user) != "alcohol_admin":
        raise Http404
    users = User.objects.all()
    context = {'users':users}
    return render(request, 'alcoholchecks/check.html', context)

@login_required
def delete_info(request, info_id):
    #記録を削除
    info = get_object_or_404(Info, id=info_id)
    owner = info.owner
    if str(request.user) != "alcohol_admin":
        raise Http404

    if request.method == 'POST':
        info.delete()
        return redirect('alcoholchecks:infos', user_id=owner.id)

    context = {'owner':owner, 'info':info}
    return render(request, 'alcoholchecks/delete_info.html', context)

@login_required
def download_or_delete(request):
    if str(request.user) != "alcohol_admin":
        raise Http404
    months = sorted(list(Info.objects.values_list("date_added__month", flat=True).distinct()))
    years = sorted(list(Info.objects.values_list("date_added__year", flat=True).distinct()))
    users = list(User.objects.all())
    users.remove(User.objects.get(id=1))
    if request.method == "POST":
        download_or_delete = request.POST.get("download_or_delete")
        month = request.POST.get("month")
        year = request.POST.get('year')
        months = list(Info.objects.filter(date_added__year=year).values_list(
                                        'date_added__month', flat=True).distinct().order_by())
        
        if download_or_delete == "download":
            response = HttpResponse(content_type="application/zip")
            zip = zipfile.ZipFile(response, "w")

            zipname = f"{year}_{month}_data.zip"
            for user in users:
                infos = Info.objects.filter(date_added__month=month, date_added__year=year,owner=user)
                output = io.BytesIO()
                filename = "{}_{}_{}_data.xlsx".format(user,year, month)
                wb = xl.Workbook()
                sheet = wb.active
                header = ["日時", "車両ナンバー", "アルコール検知", "確認者"]
                for i, j in enumerate(header):
                    sheet.cell(column=i+1, row=4, value=j)
                for index,info in enumerate(infos):
                    info.date_added += datetime.timedelta(hours=9)
                    sheet.cell(row=index+5, column=1, value=info.date_added.strftime('%Y/%m/%d %H:%M'))
                    sheet.cell(row=index+5,column=2, value=int(info.carnumber))
                    sheet.cell(row=index+5, column=3, value=info.alcohol)
                    sheet.cell(row=index+5, column=4, value="武村義治")
                sheet.title = f"{year}年{month}月"
                #罫線
                side = Side(style='thin', color='000000')
                border = Border(top=side, bottom=side, left=side, right=side)

                #フォントを設定する
                font = Font(name='メイリオ')

                #書式設定
                alignment = Alignment(horizontal='center', vertical='center')

                #columnの数繰り返す
                for col in sheet.columns:
                    max_length = 0#一番文字数が多い文字列の文字数
                    column = col[0].column_letter

                    for cell in col:#一番文字数が多い文字の文字数をmax_lengthにいれる
                        if len(str(cell.value)) > max_length:#文字数よりmax_lengthが多かったらmax_lengthを更新
                            max_length = len(str(cell.value))
                        cell.font = font #フォント
                        cell.border = border #罫線を引く
                        cell.alignment = alignment #中央に寄せる

                    sheet.column_dimensions[column].width = (max_length+2) *2

                sheet.merge_cells('A1:D3')
                sheet.cell(1, 1).value = f"{str(user)} {year}年  {month}月"
                sheet.cell(1, 1).font = xl.styles.fonts.Font(size=35)
                wb.save(output)
                zip.writestr(filename, output.getvalue())

            response['Content-Disposition'] = 'attachment; filename=%s' % zipname
            return response
        elif download_or_delete == "delete":
            for info in infos:
                info.delete()
    context = {"months":months, 'years':years}
    return render(request, 'alcoholchecks/download_or_delete.html', context)
